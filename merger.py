
from copy import copy
import os
import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.styles import Border,Side,Alignment

os.chdir(r"D:\python_project\log")
logs_path = os.getcwd()
report = os.path.join(logs_path,"report.xlsx")

def set_format(ws, row, col):
    #设置单元格格式
    thin = Side(border_style = "thin",color = "000000")
    border = Border(left = thin,right = thin, top = thin,bottom = thin)
    align = Alignment(horizontal = "center",vertical = "center")
   
    for i in range(1,row+1):
        for j in range(1, col+1):
            ws.cell(i, j).border = border
            ws.cell(i, j).alignment =align

def merge_ws():
    logs = [] #.xlsx文件

    for *dir, logs_file in os.walk(logs_path):
        for log_file in logs_file:
            if  log_file.endswith(".xlsx"):
                logs.append(log_file)
    des_file = logs[0]
    src_files = logs[1:]
    des_wb = xl.load_workbook(des_file)
    des_ws = des_wb.active
    des_cols = des_ws.max_column
    for src_file in src_files:
        src_wb = xl.load_workbook(src_file)
        src_ws = src_wb.active
        src_rows = src_ws.max_row
        src_cols = src_ws.max_column

        for i in range(src_rows):
            for j in range(src_cols):
                des_ws.column_dimensions[get_column_letter( des_cols+j+1)].width = src_ws.column_dimensions[get_column_letter(j+3)].width
                if src_ws.cell(row=i+1,column=j+3).has_style:
                    des_ws.cell(i+1, des_cols+j+1, src_ws.cell(row=i+1,column=j+3).value)
                    des_ws.cell(i+1, des_cols+j+1).font = copy(src_ws.cell(row=i+1,column=j+3).font)
                    des_ws.cell(i+1, des_cols+j+1).fill = copy(src_ws.cell(row=i+1,column=j+3).fill)
                    des_ws.cell(i+1, des_cols+j+1).alignment = copy(src_ws.cell(row=i+1,column=j+3).alignment)      
        for mcr in src_ws.merged_cells:
            cr = CellRange(mcr.coord)
            if ((cr.min_row <= 8 and cr.min_col >4) or (cr.min_row >= 9 and cr.min_col >2)):
                cr.shift(col_shift=(des_cols-2))       
                des_ws.merge_cells(cr.coord)
            
        des_cols += 4  
    set_format(des_ws, des_ws.max_row, des_ws.max_column)
    des_wb.save(report)

if __name__ == "__main__":
    merge_ws()